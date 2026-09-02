#!/usr/bin/env python3

from __future__ import annotations

import argparse
import hashlib
import json
import sys
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string


EXPECTED_HEADERS = [
    "Group", "Miles / pd", "Mile rate", "Pickup start", "Pickup end",
    "Rate zone", "Booking start", "Booking end", "1  per day", "2  per day",
    "3 - 4  per day", "5 - 7  per day", "8+ per day",
]
CHANGE_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Apply VipCars pricing recommendations to a global rate export.")
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--recommendations", required=True)
    parser.add_argument("--config", required=True)
    parser.add_argument("--report-output", required=True)
    parser.add_argument("--import-output", required=True)
    parser.add_argument("--summary-output")
    return parser.parse_args()


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def verify_baseline(workbook_path: Path, config_path: Path, config: dict[str, Any]) -> str:
    raw_manifest = config.get("baseline_manifest_file")
    if not raw_manifest:
        raise ValueError("Config is missing baseline_manifest_file.")
    manifest_path = Path(raw_manifest)
    if not manifest_path.is_absolute():
        manifest_path = config_path.parent / manifest_path
    manifest = load_json(manifest_path)
    expected = str(manifest.get("workbook_sha256", "")).lower()
    actual = sha256_file(workbook_path)
    if not expected or actual != expected:
        raise ValueError(
            f"Workbook does not match the baseline manifest: expected {expected or 'missing hash'}, got {actual}."
        )
    return actual


def parse_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for pattern in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            pass
    raise ValueError(f"Unsupported date value: {value!r}")


def validate_sheet(worksheet) -> None:
    if worksheet.max_column != len(EXPECTED_HEADERS):
        raise ValueError(f"Import worksheet must contain exactly 13 columns, found {worksheet.max_column}.")
    actual = [worksheet.cell(1, column).value for column in range(1, len(EXPECTED_HEADERS) + 1)]
    if actual != EXPECTED_HEADERS:
        raise ValueError(f"Unexpected import headers: {actual!r}")


def copy_row(worksheet, source_row: int, target_row: int, max_column: int | None = None) -> None:
    source_dimension = worksheet.row_dimensions[source_row]
    target_dimension = worksheet.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden
    target_dimension.outlineLevel = source_dimension.outlineLevel
    target_dimension.collapsed = source_dimension.collapsed
    for column in range(1, (max_column or worksheet.max_column) + 1):
        source = worksheet.cell(source_row, column)
        target = worksheet.cell(target_row, column)
        target.value = source.value
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)
        if source.hyperlink:
            target._hyperlink = copy(source.hyperlink)
        if source.comment:
            target.comment = copy(source.comment)


def expand_pickup_ranges(worksheet) -> int:
    validate_sheet(worksheet)
    original_rows = []
    for row_index in range(2, worksheet.max_row + 1):
        start = parse_date(worksheet.cell(row_index, 4).value)
        end = parse_date(worksheet.cell(row_index, 5).value)
        if end < start:
            raise ValueError(f"Pickup end precedes pickup start in row {row_index}.")
        dimension = worksheet.row_dimensions[row_index]
        original_rows.append((
            row_index,
            start,
            end,
            dimension.height,
            dimension.hidden,
            dimension.outlineLevel,
            dimension.collapsed,
        ))

    for row_index, start, end, *_ in reversed(original_rows):
        days = (end - start).days + 1
        for offset in range(1, days):
            target_row = row_index + offset
            worksheet.insert_rows(target_row)
            copy_row(worksheet, row_index, target_row)
        for offset in range(days):
            day = start + timedelta(days=offset)
            target_row = row_index + offset
            worksheet.cell(target_row, 4).value = day.strftime("%d/%m/%Y")
            worksheet.cell(target_row, 5).value = day.strftime("%d/%m/%Y")

    target_row = 2
    for _, start, end, height, hidden, outline_level, collapsed in original_rows:
        days = (end - start).days + 1
        for _ in range(days):
            dimension = worksheet.row_dimensions[target_row]
            dimension.height = height
            dimension.hidden = hidden
            dimension.outlineLevel = outline_level
            dimension.collapsed = collapsed
            target_row += 1
    return worksheet.max_row - 1


def find_band(duration: int, bands: list[dict[str, Any]]) -> dict[str, Any] | None:
    return next(
        (band for band in bands if int(band["min_days"]) <= duration <= int(band["max_days"])),
        None,
    )


def normalize_rate_zones(config: dict[str, Any]) -> list[dict[str, str]]:
    raw_rate_zones = config.get("rate_zones", [])
    if not isinstance(raw_rate_zones, list) or not raw_rate_zones:
        raise ValueError("Config is missing rate_zones.")
    rate_zones: list[dict[str, str]] = []
    locations: set[str] = set()
    codes: set[str] = set()
    for raw in raw_rate_zones:
        item = {
            "location": str(raw.get("location", "")).strip(),
            "code": str(raw.get("code", "")).strip().upper(),
            "name": str(raw.get("name", "")).strip(),
            "metroplex": str(raw.get("metroplex", "")).strip(),
        }
        location_key = item["location"].lower()
        if not all(item.values()):
            raise ValueError("Every rate zone requires location, code, name, and metroplex.")
        if location_key in locations or item["code"] in codes:
            raise ValueError(f"Duplicate rate zone mapping: {item['location']} / {item['code']}.")
        locations.add(location_key)
        codes.add(item["code"])
        rate_zones.append(item)
    return rate_zones


def expand_rate_zones(worksheet, rate_zones: list[dict[str, str]]) -> int:
    configured_codes = {item["code"] for item in rate_zones}
    existing = [
        str(worksheet.cell(row, 6).value or "").strip().upper()
        for row in range(2, worksheet.max_row + 1)
    ]
    populated = {value for value in existing if value}
    if populated:
        if any(not value for value in existing):
            raise ValueError("Rate zone column mixes populated and blank rows.")
        unknown = sorted(populated - configured_codes)
        if unknown:
            raise ValueError(f"Workbook contains unsupported rate zones: {', '.join(unknown)}.")
        missing = sorted(configured_codes - populated)
        if missing:
            raise ValueError(f"Workbook is missing configured rate zones: {', '.join(missing)}.")
        for row, code in enumerate(existing, start=2):
            worksheet.cell(row, 6).value = code
        return worksheet.max_row - 1

    original_max_row = worksheet.max_row
    max_column = worksheet.max_column
    first_zone = rate_zones[0]
    target_row = original_max_row + 1
    for source_row in range(2, original_max_row + 1):
        worksheet.cell(source_row, 6).value = first_zone["code"]
        for rate_zone in rate_zones[1:]:
            copy_row(worksheet, source_row, target_row, max_column)
            worksheet.cell(target_row, 6).value = rate_zone["code"]
            target_row += 1
    return worksheet.max_row - 1


def build_band_plans(
    recommendations: dict[str, Any],
    bands: list[dict[str, Any]],
    rate_zones: list[dict[str, str]],
) -> tuple[dict[tuple[str, str, str], dict[str, Any]], list[dict[str, Any]]]:
    decisions = recommendations.get("decisions", [])
    expected_locations = sorted(set(recommendations.get("expected_locations", [])))
    zone_by_location = {item["location"].lower(): item for item in rate_zones}
    expected_location_keys = {location.lower() for location in expected_locations}
    missing_run_locations = [
        item["location"] for item in rate_zones if item["location"].lower() not in expected_location_keys
    ]
    unsupported_run_locations = sorted(
        location for location in expected_locations if location.lower() not in zone_by_location
    )
    if missing_run_locations or unsupported_run_locations:
        details = []
        if missing_run_locations:
            details.append(f"missing: {', '.join(missing_run_locations)}")
        if unsupported_run_locations:
            details.append(f"unsupported: {', '.join(unsupported_run_locations)}")
        raise ValueError(
            "Recommendation run must cover every configured rate zone location; " + "; ".join(details) + "."
        )
    by_key: dict[tuple[str, int, str], dict[str, Any]] = {}
    candidates: set[tuple[str, str, str]] = set()
    band_by_column = {str(band["column"]): band for band in bands}

    for decision in decisions:
        pickup_date = str(decision.get("pickup_date", ""))
        duration = int(decision.get("rental_days", 0) or 0)
        location = str(decision.get("location", ""))
        rate_zone = zone_by_location.get(location.lower())
        if not rate_zone:
            raise ValueError(f"Recommendation location has no rate zone mapping: {location or 'missing'}.")
        declared_zone = str(decision.get("rate_zone") or "").strip().upper()
        if declared_zone and declared_zone != rate_zone["code"]:
            raise ValueError(
                f"Recommendation rate zone {declared_zone} does not match {location} / {rate_zone['code']}."
            )
        by_key[(pickup_date, duration, location.lower())] = decision
        band = find_band(duration, bands)
        if band:
            candidates.add((pickup_date, str(band["column"]), rate_zone["code"]))

    plans: dict[tuple[str, str, str], dict[str, Any]] = {}
    blocked: list[dict[str, Any]] = []
    zone_by_code = {item["code"]: item for item in rate_zones}
    for pickup_date, column, zone_code in sorted(candidates):
        band = band_by_column[column]
        rate_zone = zone_by_code[zone_code]
        if band.get("update_enabled", True) is False:
            blocked.append({
                "pickup_date": pickup_date,
                "duration_band": str(band["label"]),
                "column": column,
                "rate_zone": zone_code,
                "rate_zone_name": rate_zone["name"],
                "reason": "Automatic updates are disabled because the import band is open-ended.",
            })
            continue
        required_durations = range(int(band["min_days"]), int(band["max_days"]) + 1)
        missing: list[str] = []
        ratios: list[tuple[float, dict[str, Any]]] = []
        for duration in required_durations:
            location = rate_zone["location"]
            decision = by_key.get((pickup_date, duration, location.lower()))
            if not decision:
                missing.append(f"{location}/{duration}d missing")
                continue
            if decision.get("coverage_status") != "complete" or decision.get("data_quality_status") != "ok":
                missing.append(
                    f"{location}/{duration}d {decision.get('data_quality_status') or decision.get('coverage_status')}"
                )
                continue
            try:
                ratio = float(decision.get("maximum_adjustment_ratio"))
            except (TypeError, ValueError):
                missing.append(f"{location}/{duration}d invalid ratio")
                continue
            if ratio <= 0:
                missing.append(f"{location}/{duration}d invalid ratio")
                continue
            ratios.append((ratio, decision))

        if missing or not ratios:
            blocked.append({
                "pickup_date": pickup_date,
                "duration_band": str(band["label"]),
                "column": column,
                "rate_zone": zone_code,
                "rate_zone_name": rate_zone["name"],
                "reason": "; ".join(missing) if missing else "No usable decisions.",
            })
            continue
        ratio, controlling = min(ratios, key=lambda item: item[0])
        plans[(pickup_date, column, zone_code)] = {
            "ratio": ratio,
            "band": band,
            "controlling": controlling,
            "rate_zone": rate_zone,
        }
    return plans, blocked


def apply_plans(worksheet, plans, config: dict[str, Any], annotate: bool) -> list[dict[str, Any]]:
    groups = set(config.get("apply_groups", []))
    precision = int(config.get("rate_precision", 3))
    minimum_change = float(config.get("minimum_change_eur_day", 0.001))
    changes: list[dict[str, Any]] = []
    plans_by_scope: dict[tuple[str, str], list[tuple[str, dict[str, Any]]]] = {}
    for (pickup_date, column, zone_code), plan in plans.items():
        plans_by_scope.setdefault((pickup_date, zone_code), []).append((column, plan))

    for row_index in range(2, worksheet.max_row + 1):
        group = str(worksheet.cell(row_index, 1).value or "")
        if group not in groups:
            continue
        pickup_date = parse_date(worksheet.cell(row_index, 4).value).isoformat()
        zone_code = str(worksheet.cell(row_index, 6).value or "").strip().upper()
        for column, plan in plans_by_scope.get((pickup_date, zone_code), []):
            column_index = column_index_from_string(column)
            cell = worksheet.cell(row_index, column_index)
            try:
                original = float(cell.value)
            except (TypeError, ValueError):
                continue
            updated = round(original * float(plan["ratio"]), precision)
            if abs(updated - original) < minimum_change:
                continue
            cell.value = updated
            controlling = plan["controlling"]
            change = {
                "row": row_index,
                "cell": cell.coordinate,
                "group": group,
                "rate_zone": zone_code,
                "rate_zone_name": plan["rate_zone"]["name"],
                "metroplex": plan["rate_zone"]["metroplex"],
                "pickup_date": pickup_date,
                "duration_band": str(plan["band"]["label"]),
                "original_rate": original,
                "updated_rate": updated,
                "adjustment_ratio": float(plan["ratio"]),
                "controlling_location": controlling.get("location"),
                "controlling_duration_days": controlling.get("rental_days"),
                "reason": controlling.get("reason"),
            }
            changes.append(change)
            if annotate:
                cell.fill = copy(CHANGE_FILL)
                cell.comment = Comment(
                    f"VipCars recommendation: {original} -> {updated}. "
                    f"Multiplier {float(plan['ratio']):.4f}; controlling check: "
                    f"{controlling.get('location')}, {controlling.get('rental_days')} days.",
                    "VipCars scraper",
                )
    return changes


def validate_plan_targets(worksheet, plans, config: dict[str, Any]) -> None:
    groups = set(config.get("apply_groups", []))
    available: set[tuple[str, str, str, str]] = set()
    columns_by_scope: dict[tuple[str, str], set[str]] = {}
    for pickup_date, column, zone_code in plans:
        columns_by_scope.setdefault((pickup_date, zone_code), set()).add(column)
    for row_index in range(2, worksheet.max_row + 1):
        group = str(worksheet.cell(row_index, 1).value or "")
        if group not in groups:
            continue
        pickup_date = parse_date(worksheet.cell(row_index, 4).value).isoformat()
        zone_code = str(worksheet.cell(row_index, 6).value or "").strip().upper()
        for column in columns_by_scope.get((pickup_date, zone_code), set()):
            value = worksheet.cell(row_index, column_index_from_string(column)).value
            try:
                float(value)
            except (TypeError, ValueError):
                continue
            available.add((pickup_date, column, zone_code, group))

    for (pickup_date, column, zone_code), plan in plans.items():
        missing_groups = sorted(
            group for group in groups if (pickup_date, column, zone_code, group) not in available
        )
        if missing_groups:
            raise ValueError(
                "Baseline workbook does not cover planned rate target "
                f"{pickup_date} / {zone_code} / {plan['band']['label']} days for groups: {', '.join(missing_groups)}."
            )


def style_report_sheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = copy(HEADER_FILL)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_cells in worksheet.columns:
        values = [str(cell.value or "") for cell in column_cells[:200]]
        width = min(max(max((len(value) for value in values), default=0) + 2, 10), 45)
        worksheet.column_dimensions[column_cells[0].column_letter].width = width


def add_report_sheets(workbook, recommendations, changes, blocked, source_hash, expanded_rows) -> None:
    changed_sheet = workbook.create_sheet("Changed Positions")
    changed_headers = [
        "Row", "Cell", "Group", "Rate zone", "Rate zone name", "Metroplex",
        "Pickup date", "Duration band", "Original rate",
        "Updated rate", "Adjustment ratio", "Controlling location", "Controlling duration", "Reason",
    ]
    changed_sheet.append(changed_headers)
    for change in changes:
        changed_sheet.append([
            change["row"], change["cell"], change["group"], change["rate_zone"],
            change["rate_zone_name"], change["metroplex"], change["pickup_date"],
            change["duration_band"], change["original_rate"], change["updated_rate"],
            change["adjustment_ratio"], change["controlling_location"],
            change["controlling_duration_days"], change["reason"],
        ])
    style_report_sheet(changed_sheet)

    review_sheet = workbook.create_sheet("Recommendations Review")
    review_headers = [
        "Pickup date", "Duration", "Location", "Rate zone", "Rate zone name", "Metroplex",
        "Action", "Type", "Quality", "Coverage",
        "MM rank", "MM EUR/day", "Pay Now EUR", "Pay Now EUR/day", "Pay Now share",
        "Broker markup", "Broker multiplier", "MM net EUR/day", "Benchmark", "Benchmark EUR/day",
        "Target EUR/day", "Target net EUR/day", "Max multiplier", "Reason",
    ]
    review_sheet.append(review_headers)
    for decision in recommendations.get("decisions", []):
        review_sheet.append([
            decision.get("pickup_date"), decision.get("rental_days"), decision.get("location"),
            decision.get("rate_zone"), decision.get("rate_zone_name"), decision.get("metroplex"),
            decision.get("action"), decision.get("recommendation_type"), decision.get("data_quality_status"),
            decision.get("coverage_status"), decision.get("mm_rank"), decision.get("mm_rate_eur_day"),
            decision.get("pay_now_total_eur"), decision.get("pay_now_eur_day"),
            decision.get("pay_now_share_percent"), decision.get("broker_markup_percent"),
            decision.get("broker_markup_multiplier"), decision.get("mm_net_rate_eur_day"),
            decision.get("benchmark_provider"), decision.get("benchmark_rate_eur_day"),
            decision.get("site_target_rate_eur_day"), decision.get("site_target_net_rate_eur_day"),
            decision.get("maximum_adjustment_ratio"), decision.get("reason"),
        ])
    style_report_sheet(review_sheet)

    validation_sheet = workbook.create_sheet("Validation")
    validation_sheet.append(["Check", "Status", "Details"])
    validation_sheet.append(["Baseline manifest", "OK", source_hash])
    validation_sheet.append(["Expanded source rows", "OK", expanded_rows])
    validation_sheet.append(["Changed positions", "OK", len(changes)])
    validation_sheet.append(["Blocked duration bands", "REVIEW" if blocked else "OK", len(blocked)])
    for item in blocked:
        validation_sheet.append([
            f"{item['pickup_date']} / {item['rate_zone']} / {item['duration_band']} days",
            "BLOCKED",
            item["reason"],
        ])
    style_report_sheet(validation_sheet)


def prepare_workbook(source: Path, worksheet_name: str, rate_zones: list[dict[str, str]]):
    workbook = load_workbook(source)
    if worksheet_name not in workbook.sheetnames:
        raise ValueError(f"Worksheet {worksheet_name!r} was not found.")
    for sheet in list(workbook.worksheets):
        if sheet.title != worksheet_name:
            workbook.remove(sheet)
    worksheet = workbook[worksheet_name]
    expand_pickup_ranges(worksheet)
    expanded_rows = expand_rate_zones(worksheet, rate_zones)
    return workbook, worksheet, expanded_rows


def run(args: argparse.Namespace) -> dict[str, Any]:
    workbook_path = Path(args.workbook).resolve()
    recommendations_path = Path(args.recommendations).resolve()
    config_path = Path(args.config).resolve()
    report_output = Path(args.report_output).resolve()
    import_output = Path(args.import_output).resolve()
    summary_output = Path(args.summary_output).resolve() if args.summary_output else None

    config = load_json(config_path)
    source_hash = verify_baseline(workbook_path, config_path, config)
    recommendations = load_json(recommendations_path)
    bands = config.get("duration_bands", [])
    if not bands:
        raise ValueError("Config is missing duration_bands.")
    rate_zones = normalize_rate_zones(config)
    plans, blocked = build_band_plans(recommendations, bands, rate_zones)
    worksheet_name = str(config.get("worksheet", "RateGroup Export"))

    import_book, import_sheet, expanded_rows = prepare_workbook(workbook_path, worksheet_name, rate_zones)
    validate_plan_targets(import_sheet, plans, config)
    import_changes = apply_plans(import_sheet, plans, config, annotate=False)

    report_book, report_sheet, report_expanded_rows = prepare_workbook(workbook_path, worksheet_name, rate_zones)
    report_changes = apply_plans(report_sheet, plans, config, annotate=True)
    if report_expanded_rows != expanded_rows or len(report_changes) != len(import_changes):
        raise RuntimeError("Report and import workbooks diverged during generation.")
    add_report_sheets(report_book, recommendations, report_changes, blocked, source_hash, expanded_rows)

    report_output.parent.mkdir(parents=True, exist_ok=True)
    import_output.parent.mkdir(parents=True, exist_ok=True)
    import_book.save(import_output)
    report_book.save(report_output)

    summary = {
        "source_workbook_sha256": source_hash,
        "expanded_source_row_count": expanded_rows,
        "change_count": len(import_changes),
        "blocked_band_count": len(blocked),
        "rate_zone_count": len(rate_zones),
        "blocked_bands": blocked,
        "report_output": str(report_output),
        "import_output": str(import_output),
    }
    if summary_output:
        summary_output.parent.mkdir(parents=True, exist_ok=True)
        summary_output.write_text(json.dumps(summary, indent=2) + "\n", encoding="utf-8")
    return summary


def main() -> None:
    try:
        summary = run(parse_args())
        print(json.dumps(summary, indent=2))
    except Exception as error:
        print(f"VipCars rate update failed: {error}", file=sys.stderr)
        raise SystemExit(1) from error


if __name__ == "__main__":
    main()
