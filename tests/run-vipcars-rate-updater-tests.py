#!/usr/bin/env python3

from __future__ import annotations

import hashlib
import json
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "tools" / "update_vipcars_rates.py"
HEADERS = [
    "Group", "Miles / pd", "Mile rate", "Pickup start", "Pickup end",
    "Rate zone", "Booking start", "Booking end", "1  per day", "2  per day",
    "3 - 4  per day", "5 - 7  per day", "8+ per day",
]


def build_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "RateGroup Export"
    worksheet.append(HEADERS)
    worksheet.append(["CFAR", 0, None, "02/09/2026", "03/09/2026", None, None, None, 40, 20, 18, 17, 16])
    worksheet.append(["CDMR", 0, None, "02/09/2026", "03/09/2026", None, None, None, 50, 40, 38, 37, 36])
    worksheet.freeze_panes = "G2"
    worksheet["A1"].font = Font(bold=True)
    worksheet["A1"].fill = PatternFill(fill_type="solid", fgColor="E7E6E6")
    worksheet.column_dimensions["A"].width = 15
    workbook.save(path)


def write_recommendations(path: Path) -> None:
    decisions = []
    for pickup_date in ("2026-09-02", "2026-09-03"):
        for location, action, ratio in (
            ("Warsaw", "increase", 1.2),
            ("Krakow", "decrease", 0.8),
        ):
            decisions.append({
                "location": location,
                "pickup_date": pickup_date,
                "dropoff_date": "2026-09-04",
                "rental_days": 2,
                "currency": "EUR",
                "action": action,
                "recommendation_type": "top1_gap" if action == "increase" else "top1_undercut",
                "target_rank": 1,
                "reason": "test decision",
                "mm_rank": 1 if action == "increase" else 2,
                "mm_rate_eur_day": 10,
                "benchmark_provider": "Competitor",
                "benchmark_rate_eur_day": 12,
                "site_target_rate_eur_day": 12 if action == "increase" else 8,
                "maximum_adjustment_ratio": ratio,
                "data_quality_status": "ok",
                "coverage_status": "complete",
            })
    decisions.extend([
        {
            "location": location,
            "pickup_date": "2026-09-02",
            "dropoff_date": "2026-09-05",
            "rental_days": 3,
            "currency": "EUR",
            "action": "increase",
            "recommendation_type": "top1_gap",
            "target_rank": 1,
            "reason": "partial duration band",
            "mm_rank": 1,
            "mm_rate_eur_day": 10,
            "benchmark_provider": "Competitor",
            "benchmark_rate_eur_day": 12,
            "site_target_rate_eur_day": 12,
            "maximum_adjustment_ratio": 1.2,
            "data_quality_status": "ok",
            "coverage_status": "complete",
        }
        for location in ("Warsaw", "Krakow")
    ])
    decisions.extend([
        {
            "location": location,
            "pickup_date": "2026-09-02",
            "dropoff_date": "2026-09-10",
            "rental_days": 8,
            "currency": "EUR",
            "action": "decrease",
            "recommendation_type": "rank_step_undercut",
            "target_rank": 1,
            "reason": "open-ended import band",
            "mm_rank": 2,
            "mm_rate_eur_day": 10,
            "benchmark_provider": "Competitor",
            "benchmark_rate_eur_day": 9.5,
            "site_target_rate_eur_day": 9,
            "maximum_adjustment_ratio": 0.9,
            "data_quality_status": "ok",
            "coverage_status": "complete",
        }
        for location in ("Warsaw", "Krakow")
    ])
    decisions.extend([
        {
            "location": location,
            "pickup_date": "2026-09-03",
            "dropoff_date": "2026-09-08",
            "rental_days": 5,
            "currency": "EUR",
            "action": "hold",
            "recommendation_type": "none",
            "target_rank": None,
            "reason": "MM missing",
            "mm_rank": None,
            "mm_rate_eur_day": None,
            "benchmark_provider": None,
            "benchmark_rate_eur_day": None,
            "site_target_rate_eur_day": None,
            "maximum_adjustment_ratio": 1,
            "data_quality_status": "missing_mm",
            "coverage_status": "complete",
        }
        for location in ("Warsaw", "Krakow")
    ])
    path.write_text(json.dumps({
        "generated_at": "2026-09-02T00:00:00Z",
        "expected_locations": ["Krakow", "Warsaw"],
        "decisions": decisions,
    }), encoding="utf-8")


def main() -> None:
    with tempfile.TemporaryDirectory(prefix="vipcars-rate-updater-") as raw_temp:
        temp = Path(raw_temp)
        workbook_path = temp / "input.xlsx"
        recommendations_path = temp / "recommendations.json"
        manifest_path = temp / "baseline.json"
        config_path = temp / "config.json"
        report_path = temp / "vipcars-recommendations.xlsx"
        import_path = temp / "vipcars-rates-import-ready.xlsx"
        summary_path = temp / "summary.json"

        build_workbook(workbook_path)
        write_recommendations(recommendations_path)
        source_hash = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
        manifest_path.write_text(json.dumps({"workbook_sha256": source_hash}), encoding="utf-8")
        config_path.write_text(json.dumps({
            "worksheet": "RateGroup Export",
            "baseline_manifest_file": str(manifest_path),
            "apply_groups": ["CFAR"],
            "rate_precision": 3,
            "minimum_change_eur_day": 0.001,
            "duration_bands": [
                {"column": "I", "label": "1", "min_days": 1, "max_days": 1},
                {"column": "J", "label": "2", "min_days": 2, "max_days": 2},
                {"column": "K", "label": "3-4", "min_days": 3, "max_days": 4},
                {"column": "L", "label": "5-7", "min_days": 5, "max_days": 7},
                {"column": "M", "label": "8+", "min_days": 8, "max_days": 14, "update_enabled": False},
            ],
        }), encoding="utf-8")

        result = subprocess.run([
            sys.executable, str(SCRIPT),
            "--workbook", str(workbook_path),
            "--recommendations", str(recommendations_path),
            "--config", str(config_path),
            "--report-output", str(report_path),
            "--import-output", str(import_path),
            "--summary-output", str(summary_path),
        ], cwd=ROOT, capture_output=True, text=True)
        assert result.returncode == 0, result.stderr or result.stdout

        import_book = load_workbook(import_path)
        assert import_book.sheetnames == ["RateGroup Export"]
        import_sheet = import_book["RateGroup Export"]
        assert [cell.value for cell in import_sheet[1]] == HEADERS
        assert import_sheet.max_row == 5
        assert import_sheet.freeze_panes == "G2"
        assert import_sheet.column_dimensions["A"].width == 15

        rows = list(import_sheet.iter_rows(min_row=2, values_only=True))
        cfar_rows = [row for row in rows if row[0] == "CFAR"]
        cdmr_rows = [row for row in rows if row[0] == "CDMR"]
        assert [(row[3], row[4]) for row in cfar_rows] == [
            ("02/09/2026", "02/09/2026"),
            ("03/09/2026", "03/09/2026"),
        ]
        assert [row[9] for row in cfar_rows] == [16, 16]
        assert [row[10] for row in cfar_rows] == [18, 18]
        assert [row[12] for row in cfar_rows] == [16, 16]
        assert [row[9] for row in cdmr_rows] == [40, 40]

        report_book = load_workbook(report_path, data_only=False)
        assert report_book.sheetnames == [
            "RateGroup Export", "Changed Positions", "Recommendations Review", "Validation"
        ]
        assert report_book["Changed Positions"].max_row == 3
        assert report_book["Recommendations Review"].max_row == 11
        review_headers = [cell.value for cell in report_book["Recommendations Review"][1]]
        assert "Pay Now EUR" in review_headers
        assert "Broker markup" in review_headers
        assert "MM net EUR/day" in review_headers
        assert "Target net EUR/day" in review_headers

        summary = json.loads(summary_path.read_text(encoding="utf-8"))
        assert summary["source_workbook_sha256"] == source_hash
        assert summary["expanded_source_row_count"] == 4
        assert summary["change_count"] == 2
        assert summary["blocked_band_count"] == 3
        assert {item["duration_band"] for item in summary["blocked_bands"]} == {"3-4", "5-7", "8+"}

        wrong_manifest = temp / "wrong-baseline.json"
        wrong_manifest.write_text(json.dumps({"workbook_sha256": "0" * 64}), encoding="utf-8")
        wrong_config = json.loads(config_path.read_text(encoding="utf-8"))
        wrong_config["baseline_manifest_file"] = str(wrong_manifest)
        wrong_config_path = temp / "wrong-config.json"
        wrong_config_path.write_text(json.dumps(wrong_config), encoding="utf-8")
        rejected_path = temp / "must-not-exist.xlsx"
        rejected = subprocess.run([
            sys.executable, str(SCRIPT),
            "--workbook", str(workbook_path),
            "--recommendations", str(recommendations_path),
            "--config", str(wrong_config_path),
            "--report-output", str(rejected_path),
            "--import-output", str(temp / "must-not-exist-import.xlsx"),
        ], cwd=ROOT, capture_output=True, text=True)
        assert rejected.returncode != 0
        assert "baseline manifest" in (rejected.stderr + rejected.stdout).lower()
        assert not rejected_path.exists()

        uncovered_recommendations_path = temp / "uncovered-recommendations.json"
        uncovered_recommendations_path.write_text(json.dumps({
            "expected_locations": ["Krakow", "Warsaw"],
            "decisions": [
                {
                    "location": location,
                    "pickup_date": "2027-01-15",
                    "dropoff_date": "2027-01-17",
                    "rental_days": 2,
                    "action": "decrease",
                    "maximum_adjustment_ratio": 0.9,
                    "data_quality_status": "ok",
                    "coverage_status": "complete",
                }
                for location in ("Warsaw", "Krakow")
            ],
        }), encoding="utf-8")
        uncovered_report = temp / "uncovered-report.xlsx"
        uncovered_import = temp / "uncovered-import.xlsx"
        uncovered = subprocess.run([
            sys.executable, str(SCRIPT),
            "--workbook", str(workbook_path),
            "--recommendations", str(uncovered_recommendations_path),
            "--config", str(config_path),
            "--report-output", str(uncovered_report),
            "--import-output", str(uncovered_import),
        ], cwd=ROOT, capture_output=True, text=True)
        assert uncovered.returncode != 0
        assert "does not cover planned rate target" in (uncovered.stderr + uncovered.stdout).lower()
        assert not uncovered_report.exists()
        assert not uncovered_import.exists()

        extra_column_workbook = temp / "extra-column.xlsx"
        build_workbook(extra_column_workbook)
        extra_book = load_workbook(extra_column_workbook)
        extra_book["RateGroup Export"]["N1"] = "Unexpected"
        extra_book.save(extra_column_workbook)
        extra_manifest = temp / "extra-column-baseline.json"
        extra_manifest.write_text(json.dumps({
            "workbook_sha256": hashlib.sha256(extra_column_workbook.read_bytes()).hexdigest()
        }), encoding="utf-8")
        extra_config = json.loads(config_path.read_text(encoding="utf-8"))
        extra_config["baseline_manifest_file"] = str(extra_manifest)
        extra_config_path = temp / "extra-column-config.json"
        extra_config_path.write_text(json.dumps(extra_config), encoding="utf-8")
        extra_result = subprocess.run([
            sys.executable, str(SCRIPT),
            "--workbook", str(extra_column_workbook),
            "--recommendations", str(recommendations_path),
            "--config", str(extra_config_path),
            "--report-output", str(temp / "extra-column-report.xlsx"),
            "--import-output", str(temp / "extra-column-import.xlsx"),
        ], cwd=ROOT, capture_output=True, text=True)
        assert extra_result.returncode != 0
        assert "exactly 13 columns" in (extra_result.stderr + extra_result.stdout).lower()

    production_config = json.loads((ROOT / "vipcars-rate-update.config.json").read_text(encoding="utf-8"))
    production_manifest = json.loads(
        (ROOT / "input" / "vipcars-baseline-manifest.json").read_text(encoding="utf-8")
    )
    production_workbook = ROOT / "input" / "vipcars-rate-group-export.xlsx"
    assert hashlib.sha256(production_workbook.read_bytes()).hexdigest() == production_manifest["workbook_sha256"]
    production_book = load_workbook(production_workbook, read_only=True, data_only=False)
    production_sheet = production_book[production_config["worksheet"]]
    assert [cell.value for cell in next(production_sheet.iter_rows(max_row=1))] == HEADERS
    assert set(production_config["apply_groups"]) == {
        "CFAR", "CFAR1", "CFAR2", "CWAR", "CWAR1", "CWAR2", "CWAR3",
        "EDAR", "IDAR", "IDAR1", "IFAR", "IFAR1", "IFAR2", "PDAR", "PFAR",
    }

    print("All VipCars rate updater tests passed.")


if __name__ == "__main__":
    main()
